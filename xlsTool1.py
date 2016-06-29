import sys
from PyQt4.QtCore import *
from PyQt4.QtGui import *
import openpyxl
class Form(QDialog):

    def __init__(self, file_="", parent=None):
        super(Form, self).__init__(parent)
        #file_=""

        self.targetfields = [ ("Nombre",True),("Dir",True),("Dir 2 (adic)",False),("Dir 3 (adic)",False),("Col",True),("CP",True),("Estado",True) ]
        labels= ['OT','Proyecto','Descripcion','Fecha de Entrega','Remesa']
        lineEdits = dict()
        layout = QVBoxLayout()
        grid= QGridLayout()
        grid.addWidget(QLabel("Archivo"),0,0)
        self.fileLE=QLineEdit(file_)
        grid.addWidget(self.fileLE,0,1)
        button0=QPushButton("&Abrir")
        grid.addWidget(button0,0,2)
        for r,l in enumerate(labels):
            grid.addWidget(QLabel(l),r+1,0)
            lineEdits[l] = QLineEdit(l)
            lineEdits[l].selectAll()
            grid.addWidget(lineEdits[l],r+1,1)
        
        self.table = QTableWidget(2,2)
        self.table.setHorizontalHeaderLabels( [ "Campo Original", "Nuevo Campo"] )
        header = self.table.horizontalHeader()
        header.setResizeMode(QHeaderView.Stretch)
        layout.addLayout(grid)
        layout.addWidget(self.table)
        self.setLayout(layout)
        
        button1=QPushButton("&Generar")
        layout.addWidget(button1)

        self.connect(button0,SIGNAL("clicked()"),self.open_file)
        
    def open_file(self):
        dlg = QFileDialog()
        dlg.setWindowTitle( 'Seleccione archivo' )
        dlg.setViewMode( QFileDialog.Detail )
        dlg.setNameFilters( [self.tr('archivo xls (*.xls)'), self.tr('archivo xlsx (*.xlsx)'), self.tr('archivo csv (*.csv)'), self.tr('archivo txt (*.txt)'), self.tr('todos los archivos (*)')] )
        
        name = dlg.getOpenFileName(self,'Open File')
        self.fileLE.setText(name)
        if name:
            self.read_excel_file(unicode(name))

    def read_excel_file(self,file_):
        wb=openpyxl.load_workbook(file_)
        sheets=wb.get_sheet_names()
        activeSheet=wb.active
        self.table.setRowCount(activeSheet.max_column)
        print activeSheet.max_column
        for c in range(activeSheet.max_column):
            self.table.setCellWidget(c,0, QLabel(activeSheet.cell(row=1, column=c+1).value))
        self.prepare_comboboxes()

    def prepare_comboboxes(self):
        for r in range(int(self.table.rowCount)):
            cb = QComboBox()
            for f,r in self.tagetFields:
                cd.addItem(f)
            self.table.setCellWidget(0,r+1, cb)
            
                
                    
            
        
        
        
def main():
    app = QApplication(sys.argv)
    form = Form()
    form.show()
    app.exec_()


if __name__ == '__main__':
    main()
    

                
