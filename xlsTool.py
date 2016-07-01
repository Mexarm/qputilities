#!/usr/bin/python

import sys
import os.path
import csv

import openpyxl

import xlsTool_ui
              
from PyQt4 import QtGui
from PyQt4 import QtCore

def unicode_csv_reader(utf8_data, dialect=csv.excel, **kwargs):
    csv_reader = csv.reader(utf8_data, dialect=dialect, **kwargs)
    for row in csv_reader:
        #yield [unicode(cell, 'utf-8') for cell in row]
        yield [unicode(cell, 'iso-8859-1') for cell in row]
        

class xlsToolApp(QtGui.QMainWindow, xlsTool_ui.Ui_MainWindow):
    def __init__(self):
        super(self.__class__, self).__init__()
        self.setupUi(self)

        # tupla ( nombre de campo, Es necesario ? , diccionario de alias)
        self.targetFields = [ ("CUENTA",True,{'alias':["CODIGO","CODIGO DE BARRAS","BARCODE"]}),
                              ("NOMBRE",True,{'alias':[]}),
                              ("NOMBRE 2",False,{'alias':[]}),
                              ("NOMBRE 3",False,{'alias':[]}),
                              ("CALLE Y NUMERO",True,{'alias':[]}),
                              ("CALLE Y NUMERO 2",False,{'alias':[]}),
                              ("CALLE Y NUMERO 3",False,{'alias':[]}),
                              ("COLONIA",True,{'alias':["COL"]}) ,
                              ("MUNICIPIO",True,{'alias':["POBLACION"]}),
                              ("ESTADO",True,{'alias':["EDO"]}),
                              ("CODIGO_POS",True,{'alias':["CP","CODIGO_POSTAL"]}),
                              ("TELEFONO",False,{'alias':["TEL"]})]

        self.extraField = ("EXTRA",False,{'alias':[]})
        self.removeTag=   ("REMOVER CAMPO",False,{'alias':[]})
        

        self.trow=-1
        self.tcol=-1
        self.previewrecords=50
        self.table_max_rows=0
        self.valid_input_file=False
        self.valid_field_mapping=False

        self.connect(self.btn_Abrir,QtCore.SIGNAL("clicked()"),self.open_file)
        self.connect(self.btn_generar,QtCore.SIGNAL("clicked()"),self.btn_Genera_Clicked)
        self.connect(self.action_Abrir,QtCore.SIGNAL("triggered()"),self.open_file)
        self.connect(self.table,QtCore.SIGNAL("cellClicked(int,int)"),self.table_clicked)
        self.connect(self.action_Salir,QtCore.SIGNAL("triggered()"),self.closeEvent)
        

        self.progressBar.setRange(0,100)
        self.progressBar.setValue(0.0)
        self.progressBar.setTextVisible(True)
        self.progressBar.setVisible(False)
        
        #header = self.table.horizontalHeader()
        #header.setResizeMode(QtGui.QHeaderView.Stretch)
        
    def open_file(self):
        dlg = QtGui.QFileDialog()
        dlg.setWindowTitle( 'Seleccione archivo' )
        dlg.setViewMode( QtGui.QFileDialog.Detail )
        dlg.setNameFilters( [self.tr('archivo xls (*.xls)'), self.tr('archivo xlsx (*.xlsx)'), self.tr('archivo csv (*.csv)'), self.tr('archivo txt (*.txt)'), self.tr('todos los archivos (*)')] )
        
        name = dlg.getOpenFileName(self,'Open File')
        self.le_Archivo.setText(name)
        if self.is_excel_file(unicode(name)):
            self.read_excel_file(unicode(name))
        elif self.is_csv_file(unicode(name)):
            self.read_csv_file(unicode(name))
        else:
            print "No valido"
        self.apply_table_color()
        self.adjust_row_labels()
        
    def apply_table_color(self):
        if self.table.rowCount()>=2:

            r=0    
            for c in range(self.table.columnCount()):
                self.table.item(r,c).setForeground(QtGui.QColor(0,0,255))
            r=1    
            for c in range(self.table.columnCount()):
                self.table.item(r,c).setForeground(QtGui.QColor(255,0,0))

    def adjust_row_labels(self):
        #self.table.verticalHeader().setVisible(False)
        print ([ "ORIGINAL", "NUEVO"] + [str(i) for i in range(1,self.table.rowCount()-2)])
        self.table.setVerticalHeaderLabels( [ "ORIGINAL", "NUEVO"] + [str(i) for i in range(1,self.table.rowCount()-2)] )
                
    def is_csv_file(self,name):
        return (str(os.path.splitext(name)[1]).upper()) in ['.CSV','.TXT']

    def is_excel_file(self,name):
        return (str(os.path.splitext(name)[1]).upper()) in ['.XLS','.XLSX']

      
        
    def read_csv_file(self,file_):
        self.valid_input_file=False
        csv_rows=list()
        maxrows=self.previewrecords
        if self.cb_SkipFirstRow.isChecked():
            maxrows+=1
        with open(file_,'rb') as csvfile:
            reader = unicode_csv_reader(csvfile)
            
            for row in reader:
                csv_rows.append(row)
                if len(csv_rows) >= maxrows:
                    break
       
        self.table_max_rows=len(csv_rows)
            
        self.table.setRowCount(self.table_max_rows+2 )
        self.table.setColumnCount(len(csv_rows[0]))           
        for i,f in enumerate(csv_rows[0]):
            twi=QtGui.QTableWidgetItem()
            if self.cb_SkipFirstRow.isChecked():
                twi.setText(f)
                
            else:
                twi.setText("Columna {}".format(str(i+1)))
            twi.setFlags(QtCore.Qt.ItemIsEnabled)
            self.table.setItem(0,i, twi)
            self.table.setItem(1,i,QtGui.QTableWidgetItem(self.extraField[0]))

        if self.cb_SkipFirstRow.isChecked():
            csv_rows=csv_rows[1:]

        for r,row in enumerate(csv_rows):
            for c,f in enumerate(row):
                twi=QtGui.QTableWidgetItem()
                twi.setText(unicode(f))
                twi.setFlags(QtCore.Qt.ItemIsEnabled)
                self.table.setItem(r+2,c, twi)
                
        self.valid_input_file=True   
        self.table.resizeColumnsToContents()       
              

    def read_excel_file(self,file_):
        self.valid_input_file=False
        wb=openpyxl.load_workbook(file_,read_only=True, data_only=True)
        sheets=wb.get_sheet_names()
        activeSheet=wb.active
        if self.previewrecords < activeSheet.max_row:
            self.table_max_rows=self.previewrecords
        else:
            self.table_max_rows=activeSheet.max_row
        #print self.table_max_rows,activeSheet.max_column
        self.table.setRowCount(self.table_max_rows+2 )
        self.table.setColumnCount(activeSheet.max_column)
        #print activeSheet.max_column
        for c in range(1,activeSheet.max_column+1):
            twi=QtGui.QTableWidgetItem()
            if self.cb_SkipFirstRow.isChecked():
                twi.setText(activeSheet.cell(row=1, column=c).value)
                
            else:
                twi.setText("Columna {}".format(str(c)))
            twi.setFlags(QtCore.Qt.ItemIsEnabled)
            self.table.setItem(0,c-1, twi)
            self.table.setItem(1,c-1,QtGui.QTableWidgetItem(self.extraField[0]))
            
            
        self.show_preview_records(activeSheet)
        self.valid_input_file=True
        self.table.resizeColumnsToContents()  

    def show_preview_records(self,activeSheet):
        beginr=1
        endr=self.table_max_rows
        if self.cb_SkipFirstRow.isChecked():
            beginr+=1
            endr+=1
            
        for r in range (beginr,endr):
            for c in range(1,activeSheet.max_column+1):
                twi=QtGui.QTableWidgetItem()
                twi.setText(unicode(activeSheet.cell(row=r, column=c).value))
                twi.setFlags(QtCore.Qt.ItemIsEnabled)
                self.table.setItem(r,c-1, twi)

    def get_table_row(self,row):
        row_items=[]
        for col in xrange(self.table.columnCount()):
            row_items.append(self.table.item(row,col))
        return row_items
            
    def table_clicked(self,row,col):
        print "table clicked: {}, {}".format(row,col)

        if row != 1 :
            return
        if self.trow >= 0:
              
            self.table.setCellWidget(self.trow, self.tcol, None)
            
        
                
            
        self.trow = row
        self.tcol = col
        self.comb = QtGui.QComboBox()

        
        field_set=set([unicode(i.text()) for i in self.get_table_row(1) if unicode(i.text())])
        
        
        field_set = set([i[0] for i in self.targetFields])-field_set 
        field_set.add(unicode(self.extraField[0]))
        field_set.add(unicode(self.removeTag[0]))
        
        i=self.table.item(row,col)
        
        if unicode(i.text()):
            field_set.add(unicode(i.text()))
            
        for v in sorted(field_set):
            self.comb.addItem(v)
        

        
        if i:
            if self.comb.findText(unicode(i.text())):
                self.comb.setCurrentIndex(self.comb.findText(unicode(i.text())))
        self.table.setCellWidget(row, col, self.comb)
        def uf():
            return self.updatefield(row,col)
        
        self.connect(self.comb,QtCore.SIGNAL("currentIndexChanged(int)"),uf)
                     
    def updatefield(self,row,col):
        self.table.item(row,col).setText(self.comb.currentText())

    def check_field_mapping(self):
        set1=set([unicode(i) for i,j,k in self.targetFields if j])
        set2=set([unicode(i.text()) for i in self.get_table_row(1) if unicode(i.text())])
        self.valid_mapping= set1.issubset(set2)
        return self.valid_mapping
                              
    def btn_Genera_Clicked(self):
        
        if self.valid_input_file and self.check_field_mapping():
            self.generate_output()
        else:
            QtGui.QMessageBox.information(self, 'Verifique Parametros',
                                            'Verifique el archivo de entrada o parametros')
            


    def generate_output(self):
##        for row in sheet.rows:
##            for cell in row:
##                print cell.value
        pass

    def closeEvent(self,event=None): #defines the window close event
        result = QtGui.QMessageBox.question(self,
                                      "Confirmar Salida...",
                                      "Esta Seguro que desea Salir ?",
                                      QtGui.QMessageBox.Yes| QtGui.QMessageBox.No)
        if event:
            event.ignore()
        
        if result == QtGui.QMessageBox.Yes:
            if event:
                event.accept()
            else:
                QtCore.QCoreApplication.instance().quit()
                
                
        
def main():
    app = QtGui.QApplication(sys.argv)
    form = xlsToolApp()                
    form.show()                        
    sys.exit(app.exec_())              


if __name__ == '__main__':             
    main()                    
