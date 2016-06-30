import sys
from PyQt4.QtCore import *
from PyQt4.QtGui import *
import openpyxl
class Form(QDialog):

    def __init__(self, file_="", parent=None):
        super(Form, self).__init__(parent)
        #file_=""

        self.targetFields = [ ("CUENTA",True),("NOMBRE",True),("NOMBRE 2",False),("NOMBRE 3",False),("CALLE Y NUMERO",True),("CALLE Y NUMERO 2",False),("CALLE Y NUMERO 3",False),("COLONIA",True) ,("MUNICIPIO",True),("ESTADO",True),("CODIGO_POS",True),("TELEFONO",False)     ]
        labels= ['OT','Proyecto','Descripcion','Fecha de Entrega','Remesa']
        lineEdits = dict()
        layout = QVBoxLayout()
        grid= QGridLayout()
        grid.addWidget(QLabel("Archivo"),0,0)
        self.fileLE=QLineEdit(file_)
        grid.addWidget(self.fileLE,0,1)
        button0=QPushButton("&Abrir")
        grid.addWidget(button0,0,2)
        for r,l in enumerate(labels):
            grid.addWidget(QLabel(l),r+1,0)
            lineEdits[l] = QLineEdit(l)
            lineEdits[l].selectAll()
            grid.addWidget(lineEdits[l],r+1,1)
        
        self.table = QTableWidget(2,2)
        self.table.setHorizontalHeaderLabels( [ "Campo Original", "Nuevo Campo"] )
        header = self.table.horizontalHeader()
        header.setResizeMode(QHeaderView.Stretch)
        layout.addLayout(grid)
        layout.addWidget(self.table)
        self.setLayout(layout)
        
        button1=QPushButton("&Generar")
        layout.addWidget(button1)

        self.connect(button0,SIGNAL("clicked()"),self.open_file)
        self.connect(button1,SIGNAL("clicked()"),self.generate)
        self.connect(self.table,SIGNAL("cellClicked(int,int)"),self.addcomb)

        self.tr=-1
        self.tc=-1

        self.setWindowTitle("XLS Tool")

    def generate(self):
        for r in range(self.table.rowCount()):
            twi0=self.table.item(r,0)
            twi =  self.table.item(r,1)
            
            if twi:
                print "Mapeo {} -> {}".format(twi0.text(),twi.text())
                
    
    def table_clicked(self,row,col):
        print "table clicked: {}, {}".format(row,col)
        if col==1:
            if self.tr >=0:
                print "table old: {}, {}".format(self.tr,self.tc),
                twi=self.table.item(self.tr,self.tc)
                print twi
                if twi:
                    
                    #self.table.removeCellWidget(self.tr,self.tc)
                    #self.table.setCellWidget(self.tr,self.tc,None)
                    l=QLabel(twi.text())
                    #          setItem()
                    self.table.setItem(self,tr,self.tc,l)
                    print self.table.item(delf.tr,selft.tc)
            cb=QComboBox()
            for f,r in self.targetFields:
                cb.addItem(f)
            self.table.setCellWidget(row,col,cb)
            self.tr=row
            self.tc=col

    def addcomb(self,row,col):
        print "table clicked: {}, {}".format(row,col)
        print self.table.item(self.tr,self.tc)
        if self.tr >= 0:
            #text=unicode(self.table.item(self.tr,self.tc).text())[:]  
            self.table.setCellWidget(self.tr, self.tc, None)
            #
            #self.table.setItem(QTableWidgetItem(text))

            #
        
                
            
        self.tr = row
        self.tc = col
        self.comb = QComboBox()

        for l,r in self.targetFields:
            self.comb.addItem(l)

        i=self.table.item(row,col)
        if i:
            if self.comb.findText(unicode(i.text())):
                self.comb.setCurrentIndex(self.comb.findText(unicode(i.text())))
        self.table.setCellWidget(row, col, self.comb)
        def uf():
            return self.updatefield(row,col)
        
        self.connect(self.comb,SIGNAL("currentIndexChanged(int)"),uf)
                     
        #print unicode(self.table.item(self.tr,self.tc).text())

    def updatefield(self,row,col):
        print "item ({},{})={}".format(row,col,self.comb.currentText())
        self.table.item(row,col).setText(self.comb.currentText())
        
        
    def open_file(self):
        dlg = QFileDialog()
        dlg.setWindowTitle( 'Seleccione archivo' )
        dlg.setViewMode( QFileDialog.Detail )
        #dlg.setNameFilters( [self.tr('archivo xls (*.xls)'), self.tr('archivo xlsx (*.xlsx)'), self.tr('archivo csv (*.csv)'), self.tr('archivo txt (*.txt)'), self.tr('todos los archivos (*)')] )
        
        name = dlg.getOpenFileName(self,'Open File')
        self.fileLE.setText(name)
        if name:
            self.read_excel_file(unicode(name))

    def read_excel_file(self,file_):
        wb=openpyxl.load_workbook(file_)
        sheets=wb.get_sheet_names()
        activeSheet=wb.active
        self.table.setRowCount(activeSheet.max_column)
        print activeSheet.max_column
        for c in range(activeSheet.max_column):
            twi=QTableWidgetItem()
            twi.setText(activeSheet.cell(row=1, column=c+1).value)
            twi.setFlags(Qt.ItemIsEnabled)
            self.table.setItem(c,0, twi)
            self.table.setItem(c,1,QTableWidgetItem(str(c)))
        #self.prepare_comboboxes()

    def prepare_comboboxes(self):
        self.targetFieldsCBs= []
        for r in range(self.table.rowCount()):
            cb = QComboBox()
            for f,r in self.targetFields:
                cb.addItem(f)
            self.table.setCellWidget(r,1, cb)
            self.targetFieldsCBs.append(cb)
        print self.targetFieldsCBs
            
                
                    
            
        
        
        
def main():
    app = QApplication(sys.argv)
    form = Form()
    form.show()
    app.exec_()


if __name__ == '__main__':
    main()
    

                
