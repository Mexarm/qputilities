#!/usr/bin/python

import sys
import os.path
import csv

import openpyxl
import xlsxwriter


import xlsTool_ui
              
from PyQt4 import QtGui
from PyQt4 import QtCore

def unicode_csv_reader(utf8_data, dialect=csv.excel, **kwargs):
    csv_reader = csv.reader(utf8_data, dialect=dialect, **kwargs)
    for row in csv_reader:
        #yield [unicode(cell, 'utf-8') for cell in row]
        yield [unicode(cell, 'iso-8859-1') for cell in row]
        

class xlsToolApp(QtGui.QMainWindow, xlsTool_ui.Ui_MainWindow):
    def __init__(self):
        super(self.__class__, self).__init__()
        self.setupUi(self)

        # tupla ( nombre de campo, Es Requerido ? ,(grupo,orden), lista de alias)
        # grupo empieza en cero
        
        self.targetFields = [ ("CUENTA",        True,(0,1),["CODIGO","CODIGO DE BARRAS","BARCODE","IDQPN","CODBAR"]),
                              ("NOMBRE",        True,(1,1),[]),
                              ("NOMBRE 2",      False,(1,2),[]),
                              ("NOMBRE 3",      False,(1,3),[]),
                              ("CALLE Y NUMERO",True,(2,1),["CALLE","DOMICILIO"]),
                              ("CALLE 2",       False,(2,2),["NUMERO_EXTERIOR"]),
                              ("CALLE 3",       False,(2,3),["NUMERO_INTERIOR"]),
                              ("CALLE 4",       False,(2,3),[]),
                              ("CALLE 5",       False,(2,3),[]),
                              ("CALLE 6",       False,(2,3),[]),
                              ("COLONIA",       True,(3,1),["COL"]) ,
                              ("MUNICIPIO",     True,(4,1),["POBLACION","DELEGACION"]),
                              ("ESTADO",        True,(5,1),["EDO"]),
                              ("CP",            True,(6,1),["CP","CODIGO_POSTAL"]),
                              ("TELEFONO",      False,(7,1),["TEL"]),
                              ("TEL 2",         False,(7,2),[]),
                              ("TEL 3",         False,(7,3),[])]

        self.extraField = ("-EXTRA-",False,{'alias':[]})
        self.removeTag=   ("-QUITAR-",False,{'alias':[]})
        

        self.trow=-1
        self.tcol=-1
        self.previewrecords=50
        self.table_max_rows=0
        self.valid_input_file=False
        self.valid_field_mapping=False

        self.connect(self.btn_Abrir,QtCore.SIGNAL("clicked()"),self.open_file)
        self.connect(self.btn_generar,QtCore.SIGNAL("clicked()"),self.btn_Genera_Clicked)
        self.connect(self.action_Abrir,QtCore.SIGNAL("triggered()"),self.open_file)
        self.connect(self.table,QtCore.SIGNAL("cellClicked(int,int)"),self.table_clicked)
        self.connect(self.action_Salir,QtCore.SIGNAL("triggered()"),self.closeEvent)
        

        self.progressBar.setRange(0,100)
        self.progressBar.setValue(0.0)
        self.progressBar.setTextVisible(True)
        self.progressBar.setVisible(False)
        
        #header = self.table.horizontalHeader()
        #header.setResizeMode(QtGui.QHeaderView.Stretch)
        
    
    def get_transformation_dict(self):
        tdict=dict()
        targetFields_on_output = [unicode(f[0]) for f in sorted(self.targetFields,key=lambda x : x[2]) if f[2][1] == 1 ]
        targetFields_dict= dict()
        for f in self.targetFields:
            targetFields_dict[f[0]]=(f[1],f[2])
        row0=[ unicode(i.text()) if i else '' for i in self.get_table_row(0)  ]
        row1=[ unicode(i.text()) if i else '' for i in self.get_table_row(1)  ]
        extra_fields= [(row0[i],i)  for i,f in enumerate(row1) if f == unicode(self.extraField[0])]
        test=[]
        for f in row1:
            if f not in [unicode(self.extraField[0]), unicode(self.removeTag[0])]:
                test.append((f,targetFields_dict[f][1],row1.index(f)))
        test= sorted(test,key= lambda x: x[1])
        last=i
        for i,f in enumerate(targetFields_on_output):
            l2= [fld[2] for fld in test if i == fld[1][0]]
            tdict[i]=(f,l2)
            last=i
        for f in extra_fields:
            last+=1
            tdict[last]=(f[0],[f[1]])
        return tdict
            
   
    def open_file(self):
        dlg = QtGui.QFileDialog()
        dlg.setWindowTitle( 'Seleccione archivo' )
        dlg.setViewMode( QtGui.QFileDialog.Detail )
        dlg.setNameFilters( [self.tr('archivo xls (*.xls)'), self.tr('archivo xlsx (*.xlsx)'), self.tr('archivo csv (*.csv)'), self.tr('archivo txt (*.txt)'), self.tr('todos los archivos (*)')] )
        
        name = dlg.getOpenFileName(self,'Open File')
        self.le_Archivo.setText(name)
        if self.is_excel_file(unicode(name)):
            self.read_excel_file(unicode(name))
        elif self.is_csv_file(unicode(name)):
            self.read_csv_file(unicode(name))
        else:
            print "No valido"
        self.apply_table_color()
        self.adjust_row_labels()
   
    def apply_table_color(self):
        if self.table.rowCount()>=2:

            r=0    
            for c in range(self.table.columnCount()):
                self.table.item(r,c).setForeground(QtGui.QColor(0,0,255))
            r=1    
            for c in range(self.table.columnCount()):
                self.table.item(r,c).setForeground(QtGui.QColor(255,0,0))

    def adjust_row_labels(self):
        #self.table.verticalHeader().setVisible(False)
        print ([ "ORIGINAL", "NUEVO"] + [str(i) for i in range(1,self.table.rowCount()-2)])
        self.table.setVerticalHeaderLabels( [ "ORIGINAL", "NUEVO"] + [str(i) for i in range(1,self.table.rowCount()-2)] )
                
    def is_csv_file(self,name):
        return (str(os.path.splitext(name)[1]).upper()) in ['.CSV','.TXT']

    def is_excel_file(self,name):
        return (str(os.path.splitext(name)[1]).upper()) in ['.XLS','.XLSX']
        
    def read_csv_file(self,file_):
        self.valid_input_file=False
        csv_rows=list()
        maxrows=self.previewrecords
        if self.cb_SkipFirstRow.isChecked():
            maxrows+=1
        with open(file_,'rb') as csvfile:
            reader = unicode_csv_reader(csvfile)
            
            for row in reader:
                csv_rows.append(row)
                if len(csv_rows) >= maxrows:
                    break
       
        self.table_max_rows=len(csv_rows)
            
        self.table.setRowCount(self.table_max_rows+2 )
        self.table.setColumnCount(len(csv_rows[0]))           
        for i,f in enumerate(csv_rows[0]):
            twi=QtGui.QTableWidgetItem()
            f1=""
            if self.cb_SkipFirstRow.isChecked():
                f1=f
                twi.setText(f)
                
            else:
                twi.setText("Columna {}".format(str(i+1)))
            twi.setFlags(QtCore.Qt.ItemIsEnabled)
            self.table.setItem(0,i, twi)
            self.table.setItem(1,i,QtGui.QTableWidgetItem(self.get_mapped_fieldname(f1)))

        if self.cb_SkipFirstRow.isChecked():
            csv_rows=csv_rows[1:]

        for r,row in enumerate(csv_rows):
            for c,f in enumerate(row):
                twi=QtGui.QTableWidgetItem()
                twi.setText(unicode(f))
                twi.setFlags(QtCore.Qt.ItemIsEnabled)
                self.table.setItem(r+2,c, twi)
                
        self.valid_input_file=True   
        self.table.resizeColumnsToContents()       
              

    def read_excel_file(self,file_):
        self.valid_input_file=False
        wb=openpyxl.load_workbook(file_,read_only=True, data_only=True)
        sheets=wb.get_sheet_names()
        activeSheet=wb.active
        if self.previewrecords < activeSheet.max_row:
            self.table_max_rows=self.previewrecords
        else:
            self.table_max_rows=activeSheet.max_row
        #print self.table_max_rows,activeSheet.max_column
        self.table.setRowCount(self.table_max_rows+2 )
        self.table.setColumnCount(activeSheet.max_column)
        #print activeSheet.max_column
        for c in range(1,activeSheet.max_column+1):
            twi=QtGui.QTableWidgetItem()
            f=""
            if self.cb_SkipFirstRow.isChecked():
                f=activeSheet.cell(row=1, column=c).value
                twi.setText(f)
                
            else:
                twi.setText("Columna {}".format(str(c)))
            twi.setFlags(QtCore.Qt.ItemIsEnabled)
            self.table.setItem(0,c-1, twi)
            self.table.setItem(1,c-1,QtGui.QTableWidgetItem(self.get_mapped_fieldname(f)))
            
            
        self.show_preview_records(activeSheet)
        self.valid_input_file=True
        self.table.resizeColumnsToContents()  

    def show_preview_records(self,activeSheet):
        beginr=1
        endr=self.table_max_rows
        if self.cb_SkipFirstRow.isChecked():
            beginr+=1
            endr+=1
            
        for r in range (beginr,endr):
            for c in range(1,activeSheet.max_column+1):
                twi=QtGui.QTableWidgetItem()
                twi.setText(unicode(activeSheet.cell(row=r, column=c).value))
                twi.setFlags(QtCore.Qt.ItemIsEnabled)
                self.table.setItem(r,c-1, twi)

    def get_table_row(self,row):
        row_items=[]
        for col in xrange(self.table.columnCount()):
            row_items.append(self.table.item(row,col))
        return row_items
            
    def table_clicked(self,row,col):
        #print "table clicked: {}, {}".format(row,col)

        if row != 1 :
            return
        if self.trow >= 0:
            self.table.setCellWidget(self.trow, self.tcol, None)
        self.trow = row
        self.tcol = col
        self.comb = QtGui.QComboBox()
        field_set=self.get_field_set()
        field_set = set([i[0] for i in self.targetFields])-field_set 
        field_set.add(unicode(self.extraField[0]))
        field_set.add(unicode(self.removeTag[0]))
        i=self.table.item(row,col)
        if unicode(i.text()):
            field_set.add(unicode(i.text()))
        for v in sorted(field_set):
            self.comb.addItem(v)
        if i:
            if self.comb.findText(unicode(i.text())):
                self.comb.setCurrentIndex(self.comb.findText(unicode(i.text())))
        self.table.setCellWidget(row, col, self.comb)
        def uf():
            return self.updatefield(row,col)
        
        self.connect(self.comb,QtCore.SIGNAL("currentIndexChanged(int)"),uf)
                     
    def updatefield(self,row,col):
        self.table.item(row,col).setText(self.comb.currentText())

    def get_mapped_fieldname(self,fieldname):
        fn=self.extraField[0]
        fieldname=fieldname.upper()
        for i,j,g0,k in self.targetFields:
            if fieldname in [li.upper() for li in set([i]+k).difference(self.get_field_set())]:
                fn=i.upper()
                break
        return fn
    
    def get_field_set(self):
        return set([unicode(i.text()) for i in self.get_table_row(1) if i])
    
    def check_field_mapping(self):
        set1=set([unicode(i) for i,j,k,L in self.targetFields if j])
        set2=self.get_field_set()
        self.valid_mapping= set1.issubset(set2)
        return self.valid_mapping
                              
    def btn_Genera_Clicked(self):
        
        if self.valid_input_file and self.check_field_mapping():
            name=unicode(self.le_Archivo.text())
            if self.is_excel_file(name):
                self.generate_output_from_excel(name)
            elif self.is_csv_file(name):
                self.generate_output_from_csv(name)
                    
        else:
            QtGui.QMessageBox.information(self, 'Verifique Parametros',
                                            'Verifique el archivo de entrada o parametros')
    def get_suffix(self):
        suffix1=[]
        fentrega = self.de_Entrega.date()
        fecha=''
        if fentrega > QtCore.QDate(2016,1,1):
            fecha = unicode(fentrega.toString())
                            
        
        suffixes= [ ('','_'),('OT_',self.le_OT.text()),('PROY_',self.le_Proyecto.text()),('REMESA_',self.le_Remesa.text()),('FECHA_ENTREGA_',fecha),('FMT_EASY','.xlsx')]
        
        suffix1=[ s+unicode(t) for s,t in suffixes if unicode(t)]
        
        return '_'.join(suffix1)
        
    def generate_output_from_excel(self,file_):
        #Lee archivo de entrada y lo escribe en archivo de salida
        #lee xls
        if self.valid_input_file:
            
            wb=openpyxl.load_workbook(file_,use_iterators = True, data_only=True)
            #sheets=wb.get_sheet_names()
            ws=wb.active
            row_count=ws.max_row
            workbook = xlsxwriter.Workbook(unicode(os.path.splitext(file_)[0]).upper() + self.get_suffix())
            worksheet = workbook.add_worksheet()

            tdict= self.get_transformation_dict()
            
            header1=[]
            for i in range(len(tdict)):
                header1.append(tdict[i][0])
            r = 0
            for c,f in enumerate(header1):
                worksheet.write_string  (r, c,     f             )
            r = 1
            firstrow=True
            self.progressBar.setVisible(True)
            for row in ws.iter_rows():
                self.progressBar.setValue((r/float(row_count)) * 100)
                if firstrow and self.cb_SkipFirstRow.isChecked():
                    firstrow=False
                    continue
                trow=self.get_transformed_row(tdict,row, lambda x: x.value)
                
                for c,cell in enumerate(trow):
                     worksheet.write_string  (r, c,     cell              )
                r += 1
                

            workbook.close()
            self.progressBar.setValue((r/float(row_count)) * 100)
    def generate_output_from_csv(self,file_):
        #Lee archivo de entrada y lo escribe en archivo de salida
        #lee xls
        
        if self.valid_input_file:
            row_count=0
            with open(file_,'rb') as csvfile:
                reader = unicode_csv_reader(csvfile)
                row_count = sum(1 for row in reader)
            print row_count

            with open(file_,'rb') as csvfile:
                reader = unicode_csv_reader(csvfile)
 
                workbook = xlsxwriter.Workbook(unicode(os.path.splitext(file_)[0]).upper() + self.get_suffix())
                worksheet = workbook.add_worksheet()

                tdict= self.get_transformation_dict()
                
                header1=[]
                for i in range(len(tdict)):
                    header1.append(tdict[i][0])
                r = 0
                for c,f in enumerate(header1):
                    worksheet.write_string  (r, c,     f             )
                r = 1
                firstrow=True
                
                self.progressBar.setVisible(True)
                for row in reader:
                    
                    
                    
                    self.progressBar.setValue((r/float(row_count)) * 100)
                    if firstrow and self.cb_SkipFirstRow.isChecked():
                        firstrow=False
                        continue
                    trow=self.get_transformed_row(tdict,row, lambda x: x)
                    
                    for c,cell in enumerate(trow):
                         worksheet.write_string  (r, c,     cell              )
                    r += 1
                    
                    

                workbook.close()
                self.progressBar.setValue((r/float(row_count)) * 100)

    def reformat_cp(self, cp):
        fvalue=''
        
        if type(cp) is unicode:
            try:
                number = int(cp.strip().split()[0])
            except (ValueError, IndexError):
                number = None
            if number:
                fvalue = str(number).zfill(5)
            else:
                fvalue= cp
            return fvalue
        if type(cp) is int or type(cp) is long:
            fvalue = str(cp).zfill(5)
            return fvalue
        return cp                
            
            
    def get_transformed_row(self, tdict, row, fn = lambda x : x):
        trow=[]
        for i in range(len(tdict)):
            fvalues=[]
            for f in tdict[i][1]:
                if row[f]:
                    if fn(row[f]):
                        if tdict[i][0]== 'CP':
                            fvalues.append(self.reformat_cp(fn(row[f])))
                        else:
                            fvalues.append(unicode(fn(row[f])))
            fvalues = [unicode(v) for v in fvalues]
            trow.append(" ".join(fvalues))
        
        return trow
    
            
            
    def closeEvent(self,event=None): #defines the window close event
        result = QtGui.QMessageBox.question(self,
                                      "Confirmar Salida...",
                                      "Esta Seguro que desea Salir ?",
                                      QtGui.QMessageBox.Yes| QtGui.QMessageBox.No)
        if event:
            event.ignore()
        
        if result == QtGui.QMessageBox.Yes:
            if event:
                event.accept()
            else:
                QtCore.QCoreApplication.instance().quit()
                
                
        
def main():
    app = QtGui.QApplication(sys.argv)
    form = xlsToolApp()                
    form.show()                        
    sys.exit(app.exec_())              


if __name__ == '__main__':             
    main()                    
